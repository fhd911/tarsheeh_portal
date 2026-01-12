# tarsheeh/views.py
from __future__ import annotations

from decimal import Decimal, InvalidOperation
from typing import Any

from django.contrib import messages
from django.db import transaction
from django.db.models import Q
from django.forms import modelformset_factory
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse

from .forms import (
    ApplicationCreateForm,
    BatchForm,
    CandidateForm,
    OpportunityForm,
    PasteImportForm,
    UploadExcelForm,
)
from .models import Application, Batch, Candidate, Opportunity

# ======================================================
# Helpers
# ======================================================

_ARABIC_DIGITS = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")


def _to_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _to_dec(v: Any) -> Decimal:
    """
    يحوّل النص/الرقم إلى Decimal:
    - يدعم الأرقام العربية
    - يدعم الفواصل
    - لو فاضي -> 0
    """
    s = _to_str(v).translate(_ARABIC_DIGITS).replace(",", ".")
    if not s:
        return Decimal("0")
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _norm_header(h: Any) -> str:
    s = _to_str(h).replace("\ufeff", "")  # BOM
    s = " ".join(s.split())
    return s


def _pick_role(row_role: str, default_role: str) -> str:
    rr = _to_str(row_role)
    choices = dict(Application.ROLE_CHOICES)
    if rr in choices:
        return rr
    return default_role if default_role in choices else "مدير"


# ======================================================
# Excel column mapping (Upload Excel)
# ======================================================

HEADER_MAP = {
    "اسم المتقدم": "full_name",
    "السجل المدني": "national_id",
    "رقم الجوال": "phone",
    "العمل الحالي": "current_job",
    "سنوات عمل وكيل": "years_deputy",
    "سنوات عمل مدير": "years_director",
    "مدرسة المتقدم": "applicant_school",
    "قطاع المتقدم": "applicant_sector",
    "الوظيفة المتقدم عليها": "applied_role",
    "درجة الملف": "file_score",
    "درجة المقابلة": "interview_score",
    "الفرصة": "opportunity_name",
    "قطاع الفرصة": "opportunity_sector",
}

# ترتيب الأعمدة عند عدم وجود صف عناوين (A..R)
DEFAULT_ORDER = [
    "full_name",
    "national_id",
    "phone",
    None,  # التخصص (غير مستخدم)
    None,  # الرتبة (غير مستخدم)
    "current_job",
    None,  # تاريخ المباشرة (هجري) (غير مستخدم)
    "applicant_school",
    "applicant_sector",
    "applied_role",
    "opportunity_name",
    "opportunity_sector",
    None,  # سبق العمل في الإدارة المدرسية (غير مستخدم)
    "years_director",
    "years_deputy",
    "file_score",
    "interview_score",
    None,  # المجموع الكلي (نحسبه نحن)
]


def _read_xlsx_rows(file_obj) -> list[dict[str, Any]]:
    """
    - يدعم وجود صف عناوين عربية
    - وإذا لم توجد عناوين: يعتمد DEFAULT_ORDER (A..R)
    """
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise RuntimeError("openpyxl غير مثبت. نفّذ: pip install openpyxl") from e

    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_idx = None
    header_cells = None
    for i, r in enumerate(rows[:10]):
        cells = [_norm_header(c) for c in r]
        joined = " | ".join(cells)
        if ("اسم المتقدم" in joined) and ("السجل المدني" in joined):
            header_idx = i
            header_cells = cells
            break

    data: list[dict[str, Any]] = []

    if header_idx is not None and header_cells is not None:
        idx_map: dict[int, str] = {}
        for ci, raw in enumerate(header_cells):
            key = HEADER_MAP.get(raw)
            if key:
                idx_map[ci] = key

        for r in rows[header_idx + 1 :]:
            if not any(_to_str(x) for x in r):
                continue
            d: dict[str, Any] = {}
            for ci, key in idx_map.items():
                if ci < len(r):
                    d[key] = r[ci]
            data.append(d)
        return data

    for r in rows:
        if not any(_to_str(x) for x in r):
            continue
        d: dict[str, Any] = {}
        for i, key in enumerate(DEFAULT_ORDER):
            if key and i < len(r):
                d[key] = r[i]
        data.append(d)

    return data


# ======================================================
# Pages
# ======================================================

def home(request):
    """
    لتفادي NoReverseMatch في base.html (tarsheeh:home).
    اربطها في urls.py كـ name="home" أو عدّل الرابط في base.html.
    """
    return redirect("tarsheeh:dashboard")


def dashboard(request):
    stats = {
        "batches": Batch.objects.count(),
        "opps": Opportunity.objects.count(),
        "candidates": Candidate.objects.count(),
        "apps": Application.objects.count(),
        "eligible": Application.objects.filter(is_eligible=True).count(),
        "pending_scores": Application.objects.filter(Q(file_score=0) | Q(interview_score=0)).count(),
    }
    batches = Batch.objects.order_by("-created_at")[:10]
    opps = Opportunity.objects.order_by("-created_at")[:12]
    return render(request, "tarsheeh/dashboard.html", {"stats": stats, "batches": batches, "opps": opps})


def create_batch(request):
    form = BatchForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "تم إنشاء الدفعة ✅")
        return redirect("tarsheeh:dashboard")
    return render(request, "tarsheeh/form.html", {"title": "إنشاء دفعة", "form": form})


def create_opportunity(request):
    form = OpportunityForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "تم إنشاء الفرصة ✅")
        return redirect("tarsheeh:dashboard")
    return render(request, "tarsheeh/form.html", {"title": "إنشاء فرصة/مدرسة", "form": form})


@transaction.atomic
def add_manual(request):
    cand_form = CandidateForm(request.POST or None, prefix="c")
    app_form = ApplicationCreateForm(request.POST or None, prefix="a")

    if request.method == "POST" and cand_form.is_valid() and app_form.is_valid():
        national_id = cand_form.cleaned_data["national_id"]

        cand, _ = Candidate.objects.get_or_create(
            national_id=national_id,
            defaults={"full_name": cand_form.cleaned_data["full_name"]},
        )
        for f, v in cand_form.cleaned_data.items():
            setattr(cand, f, v)
        cand.save()

        app = app_form.save(commit=False)
        app.candidate = cand

        # احسب المجموع فورًا لو الدرجات موجودة
        fs = app.file_score or Decimal("0")
        iv = app.interview_score or Decimal("0")
        app.total_score = fs + iv

        app.save()

        messages.success(request, "تمت إضافة المرشح ✅")
        return redirect(reverse("tarsheeh:applications") + f"?batch={app.batch_id}&opp={app.opportunity_id}")

    return render(request, "tarsheeh/add_manual.html", {"cand_form": cand_form, "app_form": app_form})


# ======================================================
# Paste Import (Excel Paste)
# ======================================================

PASTE_HEADERS = [
    "اسم المتقدم",
    "السجل المدني",
    "رقم الجوال",
    "التخصص (اختياري)",
    "الرتبة (اختياري)",
    "العمل الحالي",
    "تاريخ المباشرة (هجري) (اختياري)",
    "مدرسة المتقدم",
    "قطاع المتقدم",
    "الوظيفة المتقدم عليها",
    "الفرصة",
    "قطاع الفرصة",
    "سبق العمل في الإدارة المدرسية (اختياري)",
    "سنوات عمل مدير",
    "سنوات عمل وكيل",
    "درجة الملف",
    "درجة المقابلة",
    "المجموع (اختياري)",
]


@transaction.atomic
def paste_import(request):
    form = PasteImportForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        batch = form.cleaned_data["batch"]
        fixed_opp = form.cleaned_data.get("opportunity")
        fixed_role = form.cleaned_data.get("applied_role") or ""

        # في الفورم عندك غالبًا اسمه text
        text = (_to_str(form.cleaned_data.get("text"))).strip()
        created = 0
        skipped = 0

        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue

            parts = [p.strip() for p in (line.split("\t") if "\t" in line else line.split(","))]
            parts += [""] * (18 - len(parts))

            full_name = parts[0]
            national_id = parts[1]
            phone = parts[2]
            current_job = parts[5]
            applicant_school = parts[7]
            applicant_sector = parts[8]
            row_role = parts[9]
            row_opp_name = parts[10]
            row_opp_sector = parts[11]
            years_director = parts[13]
            years_deputy = parts[14]
            file_score = parts[15]
            interview_score = parts[16]

            if not full_name or not national_id:
                skipped += 1
                continue

            cand, _ = Candidate.objects.get_or_create(
                national_id=national_id,
                defaults={"full_name": full_name},
            )
            cand.full_name = full_name
            cand.phone = phone
            cand.current_job = current_job
            cand.years_deputy = _to_dec(years_deputy)
            cand.years_director = _to_dec(years_director)
            cand.applicant_school = applicant_school
            cand.applicant_sector = applicant_sector
            cand.save()

            # الفرصة
            if fixed_opp:
                opp = fixed_opp
            else:
                if not row_opp_name:
                    skipped += 1
                    continue
                opp, _ = Opportunity.objects.get_or_create(
                    name=row_opp_name,
                    defaults={"sector": row_opp_sector or ""},
                )
                if row_opp_sector and opp.sector != row_opp_sector:
                    opp.sector = row_opp_sector
                    opp.save(update_fields=["sector"])

            role = fixed_role if fixed_role else row_role
            role = _pick_role(role, "مدير")

            fs = _to_dec(file_score)
            iv = _to_dec(interview_score)
            total = fs + iv

            app, is_new = Application.objects.get_or_create(
                batch=batch,
                opportunity=opp,
                candidate=cand,
                applied_role=role,
                defaults={
                    "file_score": fs,
                    "interview_score": iv,
                    "total_score": total,
                },
            )
            if not is_new:
                app.file_score = fs
                app.interview_score = iv
                app.total_score = total
                app.save(update_fields=["file_score", "interview_score", "total_score"])

            created += 1

        messages.success(request, f"تمت معالجة {created} صف ✅ (تجاوز: {skipped})")
        return redirect(reverse("tarsheeh:applications") + f"?batch={batch.id}")

    return render(
        request,
        "tarsheeh/paste_import.html",
        {
            "form": form,
            "batches": Batch.objects.order_by("-created_at"),
            "opps": Opportunity.objects.order_by("-created_at"),
            "headers": PASTE_HEADERS,  # عشان صفحة اللصق ما تطلع “سطر طويل”
        },
    )


# ======================================================
# Upload Excel (.xlsx)
# ======================================================

@transaction.atomic
def upload_excel(request):
    form = UploadExcelForm(request.POST or None, request.FILES or None)

    if request.method == "POST" and form.is_valid():
        batch: Batch = form.cleaned_data["batch"]
        excel_file = form.cleaned_data["file"]

        multi_school: bool = form.cleaned_data.get("multi_school") or False
        read_role: bool = form.cleaned_data.get("read_role_from_data") or False
        default_role: str = form.cleaned_data.get("applied_role") or "مدير"
        fixed_opp = form.cleaned_data.get("opportunity")

        if (not multi_school) and (fixed_opp is None):
            messages.error(request, "اختر مدرسة/فرصة أو فعّل (استيراد متعدد المدارس).")
            return render(request, "tarsheeh/upload_excel.html", {"form": form})

        rows = _read_xlsx_rows(excel_file)

        created_apps = 0
        updated_apps = 0
        skipped = 0

        for row in rows:
            full_name = _to_str(row.get("full_name"))
            national_id = _to_str(row.get("national_id"))
            if not full_name or not national_id:
                skipped += 1
                continue

            cand, _ = Candidate.objects.get_or_create(
                national_id=national_id,
                defaults={"full_name": full_name},
            )
            cand.full_name = full_name
            if "phone" in row:
                cand.phone = _to_str(row.get("phone"))
            if "current_job" in row:
                cand.current_job = _to_str(row.get("current_job"))
            if "years_deputy" in row:
                cand.years_deputy = _to_dec(row.get("years_deputy"))
            if "years_director" in row:
                cand.years_director = _to_dec(row.get("years_director"))
            if "applicant_school" in row:
                cand.applicant_school = _to_str(row.get("applicant_school"))
            if "applicant_sector" in row:
                cand.applicant_sector = _to_str(row.get("applicant_sector"))
            cand.save()

            # الفرصة
            if multi_school:
                opp_name = _to_str(row.get("opportunity_name"))
                opp_sector = _to_str(row.get("opportunity_sector"))
                if not opp_name:
                    skipped += 1
                    continue
                opp, _ = Opportunity.objects.get_or_create(
                    name=opp_name,
                    defaults={"sector": opp_sector or ""},
                )
                if opp_sector and opp.sector != opp_sector:
                    opp.sector = opp_sector
                    opp.save(update_fields=["sector"])
            else:
                opp = fixed_opp

            # الدور
            if read_role:
                role = _pick_role(_to_str(row.get("applied_role")), default_role)
            else:
                role = _pick_role(default_role, "مدير")

            fs = _to_dec(row.get("file_score"))
            iv = _to_dec(row.get("interview_score"))
            total = fs + iv

            app, is_new = Application.objects.get_or_create(
                batch=batch,
                opportunity=opp,
                candidate=cand,
                applied_role=role,
                defaults={"file_score": fs, "interview_score": iv, "total_score": total},
            )
            if is_new:
                created_apps += 1
            else:
                app.file_score = fs
                app.interview_score = iv
                app.total_score = total
                app.save(update_fields=["file_score", "interview_score", "total_score"])
                updated_apps += 1

        messages.success(
            request,
            f"تم الاستيراد ✅ (جديد: {created_apps} / تحديث: {updated_apps} / متجاوز: {skipped})"
        )
        return redirect(reverse("tarsheeh:applications") + f"?batch={batch.id}")

    return render(request, "tarsheeh/upload_excel.html", {"form": form})


# ======================================================
# Applications Screen
# ======================================================

def applications(request):
    batch_id = request.GET.get("batch", "")
    opp_id = request.GET.get("opp", "")
    role = request.GET.get("role", "")
    eligible = request.GET.get("eligible", "")
    q = (_to_str(request.GET.get("q"))).strip()

    qs = Application.objects.select_related("candidate", "batch", "opportunity").all()

    if batch_id:
        qs = qs.filter(batch_id=batch_id)
    if opp_id:
        qs = qs.filter(opportunity_id=opp_id)
    if role:
        qs = qs.filter(applied_role=role)
    if eligible in ("1", "0"):
        qs = qs.filter(is_eligible=(eligible == "1"))
    if q:
        qs = qs.filter(
            Q(candidate__full_name__icontains=q)
            | Q(candidate__national_id__icontains=q)
            | Q(candidate__applicant_school__icontains=q)
            | Q(opportunity__name__icontains=q)
        )

    qs = qs.order_by("-is_eligible", "-total_score", "id")

    kpi = {
        "count": qs.count(),
        "eligible": qs.filter(is_eligible=True).count(),
        "pending": qs.filter(Q(file_score=0) | Q(interview_score=0)).count(),
    }

    return render(
        request,
        "tarsheeh/applications.html",
        {
            "rows": qs[:400],
            "batches": Batch.objects.order_by("-created_at"),
            "opps": Opportunity.objects.order_by("-created_at"),
            "kpi": kpi,
            "filters": {"batch": batch_id, "opp": opp_id, "role": role, "eligible": eligible, "q": q},
        },
    )


# ======================================================
# Scores Screen (إظهار + حفظ + حساب المجموع)
# ======================================================

def scores(request):
    batch_id = request.GET.get("batch", "")
    opp_id = request.GET.get("opp", "")
    role = request.GET.get("role", "")
    pending = request.GET.get("pending", "")  # 1 = فقط النواقص

    base = Application.objects.select_related("candidate", "batch", "opportunity").all()

    if batch_id:
        base = base.filter(batch_id=batch_id)
    if opp_id:
        base = base.filter(opportunity_id=opp_id)
    if role:
        base = base.filter(applied_role=role)
    if pending == "1":
        base = base.filter(Q(file_score=0) | Q(interview_score=0))

    base = base.order_by("-is_eligible", "-total_score", "id")
    page_qs = base[:250]

    Formset = modelformset_factory(Application, fields=("file_score", "interview_score"), extra=0)

    if request.method == "POST":
        formset = Formset(request.POST, queryset=page_qs)
        if formset.is_valid():
            objs = formset.save(commit=False)
            for obj in objs:
                fs = obj.file_score or Decimal("0")
                iv = obj.interview_score or Decimal("0")
                obj.total_score = fs + iv
                obj.save(update_fields=["file_score", "interview_score", "total_score"])
            messages.success(request, "تم حفظ الدرجات وحساب المجموع ✅")
            return redirect(request.get_full_path())
        messages.error(request, "يوجد أخطاء في الإدخال. أدخل أرقام فقط.")
    else:
        formset = Formset(queryset=page_qs)

    # KPIs
    stats = {
        "total": base.count(),
        "eligible": base.filter(is_eligible=True).count(),
        "pending": base.filter(Q(file_score=0) | Q(interview_score=0)).count(),
    }

    return render(
        request,
        "tarsheeh/scores.html",
        {
            "formset": formset,
            "stats": stats,
            "batches": Batch.objects.order_by("-created_at"),
            "opps": Opportunity.objects.order_by("-created_at"),
            "filters": {"batch": batch_id, "opp": opp_id, "role": role, "pending": pending},
        },
    )


# ======================================================
# (اختياري) صفحات لو عندك روابط "عرض/تعديل" في القوالب
# ======================================================

def form(request, pk: int):
    app = get_object_or_404(Application.objects.select_related("candidate", "batch", "opportunity"), pk=pk)
    # استخدم نفس فورماتك الحالية لو عندك form.html للتعديل
    # هنا مجرد عرض سريع
    return render(request, "tarsheeh/form.html", {"title": "تعديل", "form": ApplicationCreateForm(instance=app)})


def results(request, pk: int):
    app = get_object_or_404(Application.objects.select_related("candidate", "batch", "opportunity"), pk=pk)
    return render(request, "tarsheeh/results.html", {"app": app})
